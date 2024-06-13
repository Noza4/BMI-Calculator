import os.path
from openpyxl import Workbook, load_workbook


def find_id(sheet):
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None:
                return row
    return 0


def calculate(w, h):
    h2 = h * h
    bmi = w // h2
    return bmi


def status(index):
    if index < 18.5:
        stat = "UnderWeight"
    elif 18.5 < index < 24.9:
        stat = "Normal Weight"
    elif 25 < index < 29.9:
        stat = "OverWeight"
    else:
        stat = "Obesity"
    return stat


def save_record(name, bmi, stat):
    if not os.path.exists("pydata.xlsx"):
        wb = Workbook()
    else:
        wb = load_workbook("pydata.xlsx")
    ws = wb.active
    ws.title = "Data"
    id = find_id(ws)
    ws.append([id, name, bmi, stat])
    wb.save("pydata.xlsx")

